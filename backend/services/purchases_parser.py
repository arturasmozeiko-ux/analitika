"""
Parsers for Pirkimai (Purchases) module Excel files.
"""
import re
from datetime import date
import openpyxl

# Produktų kodai, kurie visada filtruojami (pristatymai, paslaugos)
EXCLUDED_PRODUCT_CODES = {
    "001-0071",  # Prekių pristatymas
    "001-0077",  # Paletinių krovinių pristatymas
}

# Pavadinimo fragmentai, kurie filtruojami (didžiosios raidės)
EXCLUDED_NAME_PATTERNS: list[str] = []


def parse_categories_excel(file_path: str) -> list[dict]:
    """
    Parse klasių struktūra Excel.
    Returns list of {code, name, parent_code, level}.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Find header row (KODAS column)
    header_row = None
    data_start = 0
    for i, row in enumerate(rows):
        vals = [str(v).strip().upper() if v else "" for v in row]
        if "KODAS" in vals:
            header_row = vals
            data_start = i + 1
            break

    if header_row is None:
        raise ValueError("Nerasta KODAS antraštė faile")

    col_code   = next(i for i, h in enumerate(header_row) if h == "KODAS")
    col_name   = next(i for i, h in enumerate(header_row) if "PAVADINIMAS" in h)
    col_parent = next((i for i, h in enumerate(header_row) if "PAGRINDINIS" in h), None)

    def _level(code: str) -> int:
        return code.count("-")

    results = []
    seen = set()
    for row in rows[data_start:]:
        code = str(row[col_code]).strip() if row[col_code] else ""
        name = str(row[col_name]).strip() if len(row) > col_name and row[col_name] else ""
        parent = ""
        if col_parent is not None and len(row) > col_parent and row[col_parent]:
            parent = str(row[col_parent]).strip()

        # Clean non-breaking spaces
        code   = code.replace("\xa0", "").strip()
        name   = name.replace("\xa0", "").strip()
        parent = parent.replace("\xa0", "").strip()

        if not code or code == "None":
            continue
        if code in seen:
            continue
        seen.add(code)

        results.append({
            "code":        code,
            "name":        name or code,
            "parent_code": parent or None,
            "level":       _level(code),
        })

    return results


def parse_warehouse_excel(file_path: str) -> list[dict]:
    """
    Parse sandėlio kiekis Excel.
    Columns needed: Prekė, Pavadinimas, Kiekis, Savikaina, Suma,
                    Produktų grupės vadovas, Prekės kategorija
    Klasė — neprivaloma (gali nebūti naujame formate).
    Returns list of {product_code, category_code, product_name, quantity, unit_cost, total_value,
                     product_group_manager, product_category}.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Find header row — must have BOTH "Prekė" AND "Kiekis" (skip filter/metadata rows)
    header_row = None
    data_start = 0
    for i, row in enumerate(rows):
        vals = [str(v).strip().replace("\xa0", "") if v else "" for v in row]
        vals_up = [v.upper() for v in vals]
        has_preke  = any("PREK" in v for v in vals_up)
        has_kiekis = any("KIEKIS" in v for v in vals_up)
        if has_preke and has_kiekis:
            header_row = vals
            data_start = i + 1
            break

    if header_row is None:
        raise ValueError("Nerasta Prekė antraštė sandėlio faile")

    def _find(keywords):
        for kw in keywords:
            for i, h in enumerate(header_row):
                if kw.upper() in h.upper():
                    return i
        return None

    col_code    = _find(["Prekė", "Preke"])
    col_class   = _find(["Klasė", "Klase"])
    col_name    = _find(["Pavadinimas"])
    col_qty     = _find(["Kiekis"])
    col_cost    = _find(["Savikaina"])
    col_sum     = _find(["Suma"])
    col_manager = _find(["Produktų grupės vadovas", "grupės vadovas", "vadovas"])
    col_prodcat = _find(["Prekės kategorija", "kategorija"])

    if col_code is None or col_qty is None:
        raise ValueError("Nerasti reikalingi stulpeliai (Prekė, Kiekis)")

    def _float(v):
        if v is None:
            return 0.0
        try:
            return float(str(v).replace(",", ".").replace(" ", ""))
        except Exception:
            return 0.0

    results = []
    for row in rows[data_start:]:
        if not row or len(row) <= col_code:
            continue
        code = str(row[col_code]).strip().replace("\xa0", "") if row[col_code] else ""
        if not code or code == "None":
            continue
        # Skip "PAGRINDINIS" summary/total rows and footer rows (e.g. "Sugaišta:18 sekundė(s)")
        if "PAGRINDINIS" in code.upper():
            continue
        if ":" in code:  # Footer / metadata rows always contain ":"
            continue
        # Skip excluded service/delivery product codes
        if code in EXCLUDED_PRODUCT_CODES:
            continue

        cat     = str(row[col_class]).strip().replace("\xa0", "") if col_class is not None and len(row) > col_class and row[col_class] else ""
        name    = str(row[col_name]).strip() if col_name is not None and len(row) > col_name and row[col_name] else ""
        manager = str(row[col_manager]).strip() if col_manager is not None and len(row) > col_manager and row[col_manager] else ""
        prodcat = str(row[col_prodcat]).strip() if col_prodcat is not None and len(row) > col_prodcat and row[col_prodcat] else ""

        # Clean "None" strings
        if manager in ("None", "nan"): manager = ""
        if prodcat  in ("None", "nan"): prodcat  = ""

        # Also skip if name matches excluded patterns
        if any(p in name.upper() for p in EXCLUDED_NAME_PATTERNS):
            continue

        qty  = _float(row[col_qty] if len(row) > col_qty else 0)
        cost = _float(row[col_cost] if col_cost is not None and len(row) > col_cost else 0)
        suma = _float(row[col_sum]  if col_sum  is not None and len(row) > col_sum  else 0)

        results.append({
            "product_code":         code,
            "category_code":        cat,
            "product_name":         name,
            "quantity":             qty,
            "unit_cost":            cost,
            "total_value":          suma,
            "product_group_manager": manager or None,
            "product_category":      prodcat or None,
        })

    return results


def parse_sales_week_excel(file_path: str) -> list[dict]:
    """
    Parse prekių pardavimų statistika Excel.
    Columns: Prekė, Pav., [periodo stulpelis], kiekiui
    Returns list of {product_code, product_name, quantity}.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Find header row — look for row containing "kiekiui" (unique to data header, not in metadata)
    header_row = None
    data_start = 0
    for i, row in enumerate(rows):
        vals = [str(v).strip().replace("\xa0", "") if v else "" for v in row]
        vals_up = [v.upper() for v in vals]
        # "kiekiui" is unique to the real data header row of this report
        if any("KIEKIUI" in v for v in vals_up) and any("PREK" in v for v in vals_up):
            header_row = vals
            data_start = i + 1
            break

    if header_row is None:
        raise ValueError("Nerasta antraštė pardavimų faile (laukiama 'kiekiui' stulpelio)")

    def _find(keywords):
        for kw in keywords:
            for i, h in enumerate(header_row):
                if kw.upper() in h.upper():
                    return i
        return None

    col_code = _find(["Prekė", "Preke"])
    col_name = _find(["Pav"])

    # Quantity column: "kiekiui" arba paskutinis skaitinis stulpelis
    col_qty = _find(["kiekiui", "kiekis"])
    if col_qty is None:
        # Take last non-empty header column
        for i in range(len(header_row) - 1, -1, -1):
            if header_row[i].strip():
                col_qty = i
                break

    if col_code is None or col_qty is None:
        raise ValueError("Nerasti reikalingi stulpeliai (Prekė, kiekiui)")

    def _float(v):
        if v is None:
            return 0.0
        try:
            return float(str(v).replace(",", ".").replace(" ", ""))
        except Exception:
            return 0.0

    results = []
    for row in rows[data_start:]:
        if not row or len(row) <= col_code:
            continue
        code = str(row[col_code]).strip().replace("\xa0", "") if row[col_code] else ""
        if not code or code == "None":
            continue
        if "PAGRINDINIS" in code.upper():
            continue
        if code in EXCLUDED_PRODUCT_CODES:
            continue

        name = str(row[col_name]).strip() if col_name is not None and len(row) > col_name and row[col_name] else ""
        if any(p in name.upper() for p in EXCLUDED_NAME_PATTERNS):
            continue

        qty  = _float(row[col_qty] if len(row) > col_qty else 0)
        if qty <= 0:
            continue

        results.append({
            "product_code": code,
            "product_name": name,
            "quantity":     qty,
        })

    return results


# Lietuviški mėnesių pavadinimai → mėnesio numeris
MONTH_LT = {
    "SAUSIS": 1, "VASARIS": 2, "KOVAS": 3, "BALANDIS": 4,
    "GEGUŽĖ": 5, "BIRŽELIS": 6, "LIEPA": 7, "RUGPJŪTIS": 8,
    "RUGSĖJIS": 9, "SPALIS": 10, "LAPKRITIS": 11, "GRUODIS": 12,
    "GEGUZE": 5, "RUGPJUTIS": 8, "RUGSEJIS": 9,
}


def parse_annual_sales_excel(file_path: str) -> dict[int, list[dict]]:
    """
    Parse metinės pardavimų statistikos Excel (visi mėnesiai viename faile).
    Stulpeliai: Prekė, Pav., Sausis, Vasaris, ..., Gruodis, kiekiui
    Grąžina: {month_number: [{product_code, product_name, quantity}]}
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Find header row — must have "kiekiui" and "Prekė"
    header_row = None
    data_start = 0
    for i, row in enumerate(rows):
        vals = [str(v).strip().replace("\xa0", "") if v is not None else "" for v in row]
        vals_up = [v.upper() for v in vals]
        if any("KIEKIUI" in v for v in vals_up) and any("PREK" in v for v in vals_up):
            header_row = vals
            data_start = i + 1
            break

    if header_row is None:
        raise ValueError("Nerasta antraštė faile (laukiama 'kiekiui' stulpelio)")

    # Map column index → month number
    col_code = None
    col_name = None
    month_cols: dict[int, int] = {}  # {col_idx: month_num}

    for i, h in enumerate(header_row):
        h_up = h.upper().strip()
        if h_up in ("PREKĖ", "PREKE"):
            col_code = i
        elif h_up in ("PAV.", "PAV"):
            col_name = i
        elif h_up in MONTH_LT:
            month_cols[i] = MONTH_LT[h_up]

    if col_code is None or not month_cols:
        raise ValueError(f"Nerasti reikalingi stulpeliai. Rasta: {header_row}")

    def _float(v):
        if v is None:
            return 0.0
        try:
            return float(str(v).replace(",", ".").replace(" ", ""))
        except Exception:
            return 0.0

    result: dict[int, list[dict]] = {m: [] for m in range(1, 13)}

    for row in rows[data_start:]:
        if not row or len(row) <= col_code:
            continue
        code = str(row[col_code]).strip().replace("\xa0", "") if row[col_code] else ""
        if not code or code == "None":
            continue
        if "PAGRINDINIS" in code.upper():
            continue
        if code in EXCLUDED_PRODUCT_CODES:
            continue

        name = str(row[col_name]).strip() if col_name is not None and len(row) > col_name and row[col_name] else ""
        if any(p in name.upper() for p in EXCLUDED_NAME_PATTERNS):
            continue

        for col_idx, month_num in month_cols.items():
            if len(row) > col_idx:
                qty = _float(row[col_idx])
                if qty > 0:
                    result[month_num].append({
                        "product_code": code,
                        "product_name": name,
                        "quantity":     qty,
                    })


def parse_clearance_file(file_path: str) -> list[str]:
    """
    Parse išpardavimų sąrašas iš .xlsb arba .xlsx/.xls failo.
    Grąžina prekių kodų sąrašą (1-as stulpelis 'Prekė', 1-a eilutė header).
    """
    ext = file_path.lower().rsplit(".", 1)[-1]
    codes: list[str] = []

    if ext == "xlsb":
        try:
            import pyxlsb
        except ImportError:
            raise ValueError("pyxlsb biblioteka nėra įdiegta: pip install pyxlsb")
        with pyxlsb.open_workbook(file_path) as wb:
            with wb.get_sheet(1) as ws:
                for i, row in enumerate(ws.rows()):
                    if i == 0:
                        continue  # skip header
                    vals = [c.v for c in row]
                    if not vals:
                        continue
                    raw = vals[0]
                    code = str(raw).strip().replace("\xa0", "") if raw is not None else ""
                    if code and code not in ("None", "nan"):
                        codes.append(code)
    else:
        # xlsx / xls via openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        # Find header row containing "PREK"
        data_start = 1
        for i, row in enumerate(rows):
            vals_up = [str(v).strip().upper() if v else "" for v in row]
            if any("PREK" in v for v in vals_up):
                data_start = i + 1
                break
        for row in rows[data_start:]:
            if not row:
                continue
            code = str(row[0]).strip().replace("\xa0", "") if row[0] else ""
            if code and code not in ("None", "nan"):
                codes.append(code)

    return codes

    return result
